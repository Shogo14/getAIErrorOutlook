import openpyxl
import os

def write_excel(data_lists,file_name):
    sheet_name = '一覧'

    if not os.path.isfile(file_name):
        wb = openpyxl.Workbook()
        wb.create_sheet(title=sheet_name)
        wb.save(file_name)
    else:
        wb = openpyxl.load_workbook(file_name)
        if not sheet_name in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
            wb.save(file_name)
    for sheet in wb.sheetnames:
        if not sheet == sheet_name:
            wb.remove(wb[sheet])

    work_sheet = wb[sheet_name]
    max_r = work_sheet.max_row
    if max_r <= 1:
        work_sheet['A1'] = '受信日'
        work_sheet['B1'] = 'セッションID'
        work_sheet['C1'] = 'エラータイプ'
        work_sheet['D1'] = 'エラータイム'

    row_index = max_r + 1
    for data_list in data_lists:
        work_sheet.cell(row=row_index,column=1,value=data_list["received_date"])
        work_sheet.cell(row=row_index,column=2,value=data_list["session_id"])
        work_sheet.cell(row=row_index,column=3,value=data_list["errorType"])
        work_sheet.cell(row=row_index,column=4,value=data_list["errorTime"])
        row_index+=1
    wb.save(file_name)